"""
BÚSSOLA DO REPRESENTANTE v3.0
Arquivo único — sem pasta pages, sem imports externos.
SQLite para histórico incremental.
"""
import streamlit as st
import pandas as pd
import unicodedata
import re
import io
import sqlite3
from datetime import datetime
from openpyxl import load_workbook

st.set_page_config(
    page_title="Bússola do Representante",
    page_icon="🧭",
    layout="wide",
)

try:
    import plotly.graph_objects as go
    HAS_PLOTLY = True
except:
    HAS_PLOTLY = False

# ══════════════════════════════════════════════════════════════
# CSS GLOBAL
# ══════════════════════════════════════════════════════════════
st.markdown("""
<style>
section[data-testid="stSidebar"] { display:none !important; }
header[data-testid="stHeader"]   { display:none !important; }
.block-container { padding-top:0 !important; padding-bottom:2rem; max-width:1400px; }
.navbar {
    background:linear-gradient(90deg,#0D1B2A 0%,#13293D 70%,#0D1B2A 100%);
    padding:0 28px; display:flex; align-items:center;
    border-bottom:3px solid #C62828; height:56px; margin-bottom:0;
}
.navbar-brand { font-size:17px; font-weight:800; color:white; margin-right:32px; white-space:nowrap; }
.navbar-brand span { color:#C62828; }
.kpi-card { background:white; border:1px solid #E5E7EB; border-radius:14px; padding:16px 18px; }
.kpi-card.red   { border-left:4px solid #C62828; }
.kpi-card.green { border-left:4px solid #1B8A3D; }
.kpi-card.amber { border-left:4px solid #F9A825; }
.kpi-card.blue  { border-left:4px solid #1D4ED8; }
.kpi-card.navy  { border-left:4px solid #0D1B2A; }
.kpi-label { font-size:11px; color:#64748B; font-weight:700; text-transform:uppercase; letter-spacing:0.6px; margin-bottom:6px; }
.kpi-value { font-size:24px; font-weight:800; color:#0D1B2A; line-height:1.1; }
.kpi-sub   { font-size:11px; color:#94A3B8; margin-top:3px; }
.cli-card  { background:white; border:1px solid #E5E7EB; border-radius:12px; padding:12px 14px;
             margin-bottom:8px; display:flex; justify-content:space-between; align-items:center; }
.cli-card.urgente { border-left:4px solid #C62828; }
.cli-card.atencao { border-left:4px solid #F9A825; }
.cli-card.positivo{ border-left:4px solid #1B8A3D; }
.cli-name { font-size:13px; font-weight:700; color:#0D1B2A; }
.cli-sub  { font-size:11px; color:#64748B; margin-top:2px; }
.badge { font-size:11px; font-weight:700; padding:3px 10px; border-radius:20px; white-space:nowrap; }
.badge-red   { background:#FEF2F2; color:#C62828; }
.badge-amber { background:#FFFBEB; color:#B45309; }
.badge-green { background:#F0FDF4; color:#1B8A3D; }
.sec-title { font-size:14px; font-weight:700; color:#0D1B2A; margin:20px 0 10px;
             padding-bottom:6px; border-bottom:2px solid #E5E7EB; }
.app-header { background:linear-gradient(90deg,#0D1B2A 0%,#13293D 60%,#C62828 100%);
              padding:20px 28px; border-radius:16px; margin-bottom:20px; color:white; }
.app-header h1 { color:white; margin-bottom:4px; font-size:26px; }
.app-header p  { color:#E5E7EB; font-size:13px; margin:0; }
.section-card  { background:white; border:1px solid #E5E7EB; border-radius:16px;
                 padding:18px; margin-bottom:18px; }
div[data-testid="stMetricLabel"] { color:#475569; font-weight:600; }
div[data-testid="stMetricValue"] { color:#0D1B2A; font-weight:800; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# FUNÇÕES UTILITÁRIAS
# ══════════════════════════════════════════════════════════════
def _ln(t):
    t = str(t).strip().upper()
    return unicodedata.normalize("NFKD",t).encode("ASCII","ignore").decode("utf-8")

def _norm(t):
    t = str(t).strip().lower()
    for k,v in {"ç":"c","ã":"a","á":"a","à":"a","â":"a","é":"e","ê":"e",
                "í":"i","ó":"o","ô":"o","õ":"o","ú":"u"}.items():
        t=t.replace(k,v)
    return t

def _num(v):
    if pd.isna(v): return 0.0
    if isinstance(v,(int,float)): return float(v)
    v=str(v).strip().replace("R$","").replace("%","").replace(".","").replace(",",".")
    try: return float(v)
    except: return 0.0

def _m(v):
    try: return f"R$ {float(v):,.0f}".replace(",",".")
    except: return "R$ 0"

def _p(v):
    try: return f"{float(v)*100:,.1f}%".replace(",","X").replace(".",",").replace("X",".")
    except: return "0,0%"

def _ac(df, ops, exata=False):
    for o in ops:
        for c in df.columns:
            if _ln(c)==_ln(o): return c
    if not exata:
        for o in ops:
            for c in df.columns:
                if _ln(o) in _ln(c): return c
    return None

def _fc(cols, names):
    mp={_norm(c):c for c in cols}
    for n in names:
        nn=_norm(n)
        for cn,co in mp.items():
            if nn==cn: return co
        for cn,co in mp.items():
            if nn in cn: return co
    return None

def _prio(row):
    if row["GAP"]>=0: return "🟢 AUMENTAR VENDA"
    if row["PERC_GAP"]<=-0.20: return "🔴 URGENTE"
    return "🟡 ATENÇÃO"

def _acao(row):
    if row["GAP"]>=0: return "Aumentar venda"
    if row["PERC_GAP"]<=-0.20: return "Recuperar urgente"
    return "Acompanhar"

def _cp(v):
    if "URGENTE"  in str(v): return "background-color:#F8D7DA;color:#842029;font-weight:bold"
    if "ATENÇÃO"  in str(v): return "background-color:#FFF3CD;color:#664D03;font-weight:bold"
    if "AUMENTAR" in str(v): return "background-color:#D1E7DD;color:#0F5132;font-weight:bold"
    return ""

def _cn(v):
    try:
        x=float(v)
        if x<0: return "color:#C62828;font-weight:bold"
        if x>0: return "color:#1B8A3D;font-weight:bold"
    except: pass
    return ""

def _leia(s):
    return s.astype(str).str.strip().str.replace(".0","",regex=False).str.replace(" ","",regex=False).str.replace("-","",regex=False)

def _conv(s):
    if pd.api.types.is_numeric_dtype(s): return pd.to_numeric(s,errors="coerce")
    s2=s.astype(str).str.strip().str.replace("R$","",regex=False).str.replace(" ","",regex=False)
    tv=s2.str.contains(",",regex=False,na=False)
    s2.loc[tv]=s2.loc[tv].str.replace(".",",",regex=False).str.replace(",",".",regex=False)
    return pd.to_numeric(s2,errors="coerce")

def _fmt(v):
    if pd.isna(v) or v=="": return ""
    try: return f"{float(v):.2f}".replace(".",",")
    except: return ""

def _xl(f, s=0):
    f.seek(0); n=f.name.lower()
    if n.endswith(".xlsx"): return pd.read_excel(f,engine="openpyxl",sheet_name=s,header=None)
    if n.endswith(".xls"):  return pd.read_excel(f,engine="xlrd",sheet_name=s,header=None)
    return pd.read_excel(f,sheet_name=s,header=None)

def _det_h(db, words, lim=25):
    bl,bs=0,-1
    for i in range(min(len(db),lim)):
        row=[_norm(x) for x in db.iloc[i].fillna("").astype(str).tolist()]
        sc=sum(1 for c in row for w in words if w in c)
        if sc>bs: bs=sc; bl=i
    return bl

def _mk(db, h):
    cab=db.iloc[h].fillna("").astype(str).tolist()
    d=db.iloc[h+1:].copy().reset_index(drop=True); d.columns=cab; return d

# ══════════════════════════════════════════════════════════════
# BANCO DE DADOS — SQLite histórico incremental
# ══════════════════════════════════════════════════════════════
DB = "bussola.db"

def _conn(): return sqlite3.connect(DB, check_same_thread=False)

def _init_db():
    c = _conn()
    c.execute("""CREATE TABLE IF NOT EXISTS importacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo TEXT, arquivo TEXT, importado_em TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS vendas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        importacao_id INTEGER, cliente TEXT, cidade TEXT, bandeira TEXT,
        meta REAL, realizado REAL, projecao REAL, gap REAL, perc_gap REAL,
        crescimento REAL, valor_2025 REAL, prioridade TEXT, acao TEXT,
        endereco TEXT, bairro TEXT, telefone TEXT, supervisor TEXT, rota TEXT,
        importado_em TEXT)""")
    c.commit(); c.close()

def _reg_imp(tipo, arq):
    cn=_conn(); agora=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur=cn.cursor(); cur.execute("INSERT INTO importacoes (tipo,arquivo,importado_em) VALUES (?,?,?)",(tipo,arq,agora))
    iid=cur.lastrowid; cn.commit(); cn.close(); return iid, agora

def _salvar_vendas(df, iid, agora):
    cn=_conn()
    rows=[(iid, str(r.get("CLIENTE_FINAL","")), str(r.get("CIDADE","")),
           str(r.get("BANDEIRA_FINAL","")), float(r.get("META",0)),
           float(r.get("REALIZADO",0)), float(r.get("PROJECAO",0)),
           float(r.get("GAP",0)), float(r.get("PERC_GAP",0)),
           float(r.get("CRESCIMENTO",0)), float(r.get("VALOR_2025",0)),
           str(r.get("PRIORIDADE","")), str(r.get("ACAO","")),
           str(r.get("ENDERECO","")), str(r.get("BAIRRO","")),
           str(r.get("TELEFONE","")), str(r.get("SUPERVISOR_CMK","")),
           str(r.get("ROTA","")), agora)
          for _,r in df.iterrows()]
    cn.executemany("""INSERT INTO vendas
        (importacao_id,cliente,cidade,bandeira,meta,realizado,projecao,gap,perc_gap,
         crescimento,valor_2025,prioridade,acao,endereco,bairro,telefone,supervisor,rota,importado_em)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", rows)
    cn.commit(); cn.close()

def _ler_vendas():
    try:
        cn=_conn()
        df=pd.read_sql("""SELECT v.* FROM vendas v
            INNER JOIN (SELECT MAX(id) max_id FROM importacoes WHERE tipo='vendas') i
            ON v.importacao_id=i.max_id""", cn)
        cn.close()
        if df.empty: return None
        df=df.rename(columns={"cliente":"CLIENTE_FINAL","cidade":"CIDADE","bandeira":"BANDEIRA_FINAL",
            "meta":"META","realizado":"REALIZADO","projecao":"PROJECAO","gap":"GAP",
            "perc_gap":"PERC_GAP","crescimento":"CRESCIMENTO","valor_2025":"VALOR_2025",
            "prioridade":"PRIORIDADE","acao":"ACAO","endereco":"ENDERECO","bairro":"BAIRRO",
            "telefone":"TELEFONE","supervisor":"SUPERVISOR_CMK","rota":"ROTA"})
        return df
    except: return None

def _ler_hist_cliente(nome):
    try:
        cn=_conn()
        df=pd.read_sql("SELECT * FROM vendas WHERE cliente LIKE ? ORDER BY importado_em ASC",
                       cn,params=(f"%{nome}%",))
        cn.close(); return df
    except: return pd.DataFrame()

def _ler_imps():
    try:
        cn=_conn(); df=pd.read_sql("SELECT * FROM importacoes ORDER BY id DESC LIMIT 20",cn)
        cn.close(); return df
    except: return pd.DataFrame()

def _tem_dados():
    try:
        cn=_conn(); cur=cn.cursor()
        cur.execute("SELECT COUNT(*) FROM vendas"); n=cur.fetchone()[0]
        cn.close(); return n>0
    except: return False

_init_db()

# ══════════════════════════════════════════════════════════════
# PROCESSAMENTO DE CURVA + CMK
# ══════════════════════════════════════════════════════════════
def _processar_curva(curva_f, cmk_f):
    try:    curva=pd.read_excel(curva_f,sheet_name="DADOS")
    except: curva=pd.read_excel(curva_f)
    cmk=pd.read_excel(cmk_f)
    curva.columns=[_ln(c) for c in curva.columns]
    cmk.columns  =[_ln(c) for c in cmk.columns]

    def col(ops,ex=False): return _ac(curva,ops,exata=ex)
    cl=col(["CLIENTE"],ex=True); cb=col(["BANDEIRA"],ex=True)
    cc=col(["CNPJ"],ex=True);    cm=_ac(cmk,["CNPJ"],exata=True)
    cme=col(["META"],ex=True);   cre=col(["REAL"],ex=True)
    cpr=col(["REAL PROJ","REAL PROJ AC"])
    cga=col(["DESVIO PROJ"],ex=True); cpe=col(["% DESVIO"],ex=True)
    c25=col(["2025"],ex=True);   ccr=col(["% CRESC PROJ","% CRESC PROJ AC","CRESC"])

    df=curva.copy()
    df["META"]        =df[cme].apply(_num) if cme else 0.0
    df["REALIZADO"]   =df[cre].apply(_num) if cre else 0.0
    df["PROJECAO"]    =df[cpr].apply(_num) if cpr else 0.0
    df["GAP"]         =df[cga].apply(_num) if cga else df["PROJECAO"]-df["META"]
    df["PERC_GAP"]    =df[cpe].apply(_num) if cpe else 0.0
    df["VALOR_2025"]  =df[c25].apply(_num) if c25 else 0.0
    df["CRESCIMENTO"] =df[ccr].apply(_num) if ccr else 0.0
    df["CLIENTE_FINAL"] =df[cl].astype(str).str.strip() if cl else "SEM NOME"
    df["BANDEIRA_FINAL"]=df[cb].astype(str).str.strip() if cb else ""

    if cc and cm:
        df["_CNPJ_"] =df[cc].apply(lambda v:re.sub(r"\D","",str(v)))
        cmk["_CNPJ_"]=cmk[cm].apply(lambda v:re.sub(r"\D","",str(v)))
        cid=_ac(cmk,["MUNICIPIO","CIDADE","CITY"]); end=_ac(cmk,["ENDERECO","ENDEREÇO","END","LOGRADOURO"])
        bai=_ac(cmk,["BAIRRO"]); tel=_ac(cmk,["TELEFONE","FONE","TEL"])
        spv=_ac(cmk,["SUPERVISOR"]); rot=_ac(cmk,["ROTA"])
        mp={"_CNPJ_":"_CNPJ_"}
        if cid: mp[cid]="CIDADE"
        if end: mp[end]="ENDERECO"
        if bai: mp[bai]="BAIRRO"
        if tel: mp[tel]="TELEFONE"
        if spv: mp[spv]="SUPERVISOR_CMK"
        if rot: mp[rot]="ROTA"
        cmk_m=cmk[list(mp.keys())].drop_duplicates("_CNPJ_").rename(columns=mp)
        df=df.merge(cmk_m,on="_CNPJ_",how="left")

    for c2 in ["CIDADE","ENDERECO","BAIRRO","TELEFONE","SUPERVISOR_CMK","ROTA"]:
        if c2 not in df.columns: df[c2]=""

    df["PRIORIDADE"]=df.apply(_prio,axis=1)
    df["ACAO"]      =df.apply(_acao,axis=1)
    ordem={"🔴 URGENTE":1,"🟡 ATENÇÃO":2,"🟢 AUMENTAR VENDA":3}
    df["ORDEM"]=df["PRIORIDADE"].map(ordem).fillna(9)
    return df.sort_values(["ORDEM","GAP"],ascending=[True,True])

# ══════════════════════════════════════════════════════════════
# NAVBAR + ESTADO
# ══════════════════════════════════════════════════════════════
if "pagina" not in st.session_state: st.session_state["pagina"]="dashboard"

st.markdown('<div class="navbar"><div class="navbar-brand">🧭 Bússola do <span>Representante</span></div></div>',unsafe_allow_html=True)

ABAS=[("dashboard","🧭 Dashboard"),("cliente","👤 Cliente a Cliente"),
      ("cotabot","💰 CotaBot"),("cobranca","🧾 Cobrança"),("visitas","📍 Painel de Visitas")]

cols=st.columns([2.2]+[1.4]*len(ABAS))
for i,(key,label) in enumerate(ABAS):
    with cols[i+1]:
        if st.button(label,use_container_width=True,
                     type="primary" if st.session_state["pagina"]==key else "secondary",
                     key=f"nav_{key}"):
            st.session_state["pagina"]=key; st.rerun()

st.markdown("<div style='margin-top:12px;'></div>",unsafe_allow_html=True)
PAG=st.session_state["pagina"]

# ══════════════════════════════════════════════════════════════
# UPLOAD CENTRAL (compartilhado entre Dashboard e outras abas)
# ══════════════════════════════════════════════════════════════
def _bloco_upload():
    with st.expander("📁 Carregar dados do setor",expanded="df" not in st.session_state):
        u1,u2=st.columns(2)
        with u1: cf=st.file_uploader("📊 Curva Semanal (.xlsx)",type=["xlsx","xls"],key="up_curva")
        with u2: mf=st.file_uploader("📍 CMK (.xlsx)",type=["xlsx","xls"],key="up_cmk")
        if cf and mf:
            with st.spinner("Processando e salvando no histórico..."):
                try:
                    df_n=_processar_curva(cf,mf)
                    iid,agora=_reg_imp("vendas",cf.name)
                    _salvar_vendas(df_n,iid,agora)
                    st.session_state["df"]=df_n
                    st.success(f"✅ {len(df_n)} clientes carregados e salvos no histórico!")
                except Exception as e:
                    st.error(f"Erro: {e}")
        if "df" not in st.session_state and _tem_dados():
            df_b=_ler_vendas()
            if df_b is not None:
                st.session_state["df"]=df_b
                st.info("📂 Dados carregados do histórico.")
        if _tem_dados():
            with st.expander("📜 Histórico de importações"):
                st.dataframe(_ler_imps(),use_container_width=True,hide_index=True)

def _sem_dados():
    st.markdown("""
    <div style="background:#F8FAFC;border:2px dashed #CBD5E1;border-radius:16px;
                padding:60px;text-align:center;margin-top:16px;">
        <div style="font-size:52px;">🧭</div>
        <div style="font-size:22px;font-weight:800;color:#0D1B2A;margin:10px 0 6px;">Bússola do Representante</div>
        <div style="font-size:14px;color:#64748B;">
            Abra o painel acima e suba a <strong>Curva Semanal</strong> + <strong>CMK</strong>.
        </div>
    </div>""",unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════
if PAG=="dashboard":
    _bloco_upload()
    if "df" not in st.session_state: _sem_dados(); st.stop()
    df=st.session_state["df"]

    urg =int((df["PRIORIDADE"]=="🔴 URGENTE").sum())
    ate =int((df["PRIORIDADE"]=="🟡 ATENÇÃO").sum())
    pos =int((df["PRIORIDADE"]=="🟢 AUMENTAR VENDA").sum())
    mt  =df["META"].sum(); rt=df["REALIZADO"].sum()
    prt =df["PROJECAO"].sum(); gt=df["GAP"].sum()
    pct =(rt/mt*100) if mt>0 else 0

    k1,k2,k3,k4,k5=st.columns(5)
    with k1: st.markdown(f'<div class="kpi-card navy"><div class="kpi-label">Meta total</div><div class="kpi-value">{_m(mt)}</div><div class="kpi-sub">{len(df)} clientes</div></div>',unsafe_allow_html=True)
    with k2: st.markdown(f'<div class="kpi-card green"><div class="kpi-label">Realizado</div><div class="kpi-value">{_m(rt)}</div><div class="kpi-sub">{pct:.1f}% da meta</div></div>',unsafe_allow_html=True)
    with k3: st.markdown(f'<div class="kpi-card blue"><div class="kpi-label">Projeção</div><div class="kpi-value">{_m(prt)}</div><div class="kpi-sub">até fim do mês</div></div>',unsafe_allow_html=True)
    cg="red" if gt<0 else "green"
    with k4: st.markdown(f'<div class="kpi-card {cg}"><div class="kpi-label">GAP total</div><div class="kpi-value">{_m(gt)}</div><div class="kpi-sub">{"déficit" if gt<0 else "superávit"}</div></div>',unsafe_allow_html=True)
    with k5: st.markdown(f'<div class="kpi-card red"><div class="kpi-label">🔴 Urgentes</div><div class="kpi-value">{urg}</div><div class="kpi-sub">{ate} atenção · {pos} positivos</div></div>',unsafe_allow_html=True)
    st.markdown("<div style='margin-top:10px;'></div>",unsafe_allow_html=True)

    if HAS_PLOTLY:
        g1,g2=st.columns([1.4,1])
        with g1:
            st.markdown("<div class='sec-title'>📊 Faturamento do mês vs meta</div>",unsafe_allow_html=True)
            fig=go.Figure()
            fig.add_bar(name="Meta",     x=["Meta"],     y=[mt],  marker_color="#0D1B2A")
            fig.add_bar(name="Realizado",x=["Realizado"],y=[rt],  marker_color="#1B8A3D")
            fig.add_bar(name="Projeção", x=["Projeção"], y=[prt], marker_color="#1D4ED8")
            fig.update_layout(height=250,margin=dict(l=0,r=0,t=8,b=0),
                plot_bgcolor="white",paper_bgcolor="white",showlegend=True,
                legend=dict(orientation="h",y=-0.2),bargap=0.35,
                yaxis=dict(tickprefix="R$ ",tickformat=",.0f",gridcolor="#F1F5F9"))
            st.plotly_chart(fig,use_container_width=True)
        with g2:
            st.markdown("<div class='sec-title'>🎯 Prioridades do setor</div>",unsafe_allow_html=True)
            fig2=go.Figure(go.Pie(labels=["🔴 Urgente","🟡 Atenção","🟢 Positivo"],
                values=[urg,ate,pos],marker_colors=["#C62828","#F9A825","#1B8A3D"],
                hole=0.48,textinfo="label+percent",textfont_size=12))
            fig2.update_layout(height=250,margin=dict(l=0,r=0,t=8,b=0),paper_bgcolor="white",showlegend=False)
            st.plotly_chart(fig2,use_container_width=True)

        g3,g4=st.columns([1,1.4])
        with g3:
            st.markdown("<div class='sec-title'>📍 GAP por cidade</div>",unsafe_allow_html=True)
            rc=df.groupby("CIDADE").agg(GAP_TOTAL=("GAP","sum")).reset_index()
            rc=rc[rc["CIDADE"].astype(str).str.strip()!=""].sort_values("GAP_TOTAL").tail(12)
            cores=["#C62828" if v<0 else "#1B8A3D" for v in rc["GAP_TOTAL"]]
            fig3=go.Figure(go.Bar(x=rc["GAP_TOTAL"],y=rc["CIDADE"],orientation="h",
                marker_color=cores,text=[_m(v) for v in rc["GAP_TOTAL"]],
                textposition="outside",textfont=dict(size=10)))
            fig3.update_layout(height=320,margin=dict(l=0,r=80,t=8,b=0),
                plot_bgcolor="white",paper_bgcolor="white",
                xaxis=dict(tickprefix="R$ ",tickformat=",.0f",gridcolor="#F1F5F9"),
                yaxis=dict(tickfont=dict(size=11)))
            st.plotly_chart(fig3,use_container_width=True)
        with g4:
            st.markdown("<div class='sec-title'>📈 Real vs projeção — top 10</div>",unsafe_allow_html=True)
            top=df.nlargest(10,"META")[["CLIENTE_FINAL","REALIZADO","PROJECAO"]].copy()
            top["CLIENTE_FINAL"]=top["CLIENTE_FINAL"].str[:18]
            fig4=go.Figure()
            fig4.add_bar(name="Realizado",x=top["CLIENTE_FINAL"],y=top["REALIZADO"],marker_color="#1B8A3D")
            fig4.add_bar(name="Projeção", x=top["CLIENTE_FINAL"],y=top["PROJECAO"], marker_color="#1D4ED8",opacity=0.75)
            fig4.update_layout(height=320,margin=dict(l=0,r=0,t=8,b=70),
                plot_bgcolor="white",paper_bgcolor="white",barmode="group",
                showlegend=True,legend=dict(orientation="h",y=-0.35),
                yaxis=dict(tickprefix="R$ ",tickformat=",.0f",gridcolor="#F1F5F9"),
                xaxis=dict(tickangle=-35,tickfont=dict(size=10)))
            st.plotly_chart(fig4,use_container_width=True)

    # Top 5 urgentes
    st.markdown("<div class='sec-title'>🔴 Top 5 urgentes — ação imediata</div>",unsafe_allow_html=True)
    df_urg=df[df["PRIORIDADE"]=="🔴 URGENTE"].head(5)
    if df_urg.empty: st.success("Nenhum cliente urgente!")
    else:
        cu=st.columns(min(len(df_urg),5))
        for i,(_,row) in enumerate(df_urg.iterrows()):
            with cu[i]:
                st.markdown(f"""<div style="background:#FEF2F2;border:1px solid #C62828;border-radius:12px;
                    padding:14px;text-align:center;">
                    <div style="font-size:11px;font-weight:700;color:#94A3B8;">#{i+1} URGENTE</div>
                    <div style="font-size:13px;font-weight:800;color:#0D1B2A;margin:6px 0 4px;line-height:1.3;">{row["CLIENTE_FINAL"][:22]}</div>
                    <div style="font-size:11px;color:#64748B;">{row.get("CIDADE","")}</div>
                    <div style="font-size:15px;font-weight:700;color:#C62828;margin-top:6px;">{_m(row["GAP"])}</div>
                    <div style="font-size:10px;color:#94A3B8;">{_p(row["PERC_GAP"])} do GAP</div>
                    <div style="font-size:11px;font-weight:600;color:#C62828;margin-top:4px;">{row["ACAO"]}</div>
                </div>""",unsafe_allow_html=True)

    # Atenção + crescimento
    p1,p2=st.columns(2)
    with p1:
        st.markdown("<div class='sec-title'>🟡 Em atenção</div>",unsafe_allow_html=True)
        for _,row in df[df["PRIORIDADE"]=="🟡 ATENÇÃO"].head(8).iterrows():
            st.markdown(f"""<div class="cli-card atencao"><div>
                <div class="cli-name">{row["CLIENTE_FINAL"]}</div>
                <div class="cli-sub">{row.get("CIDADE","")} · GAP: {_m(row["GAP"])} · {_p(row["PERC_GAP"])}</div>
            </div><span class="badge badge-amber">Atenção</span></div>""",unsafe_allow_html=True)
    with p2:
        st.markdown("<div class='sec-title'>🟢 Maiores em crescimento</div>",unsafe_allow_html=True)
        for _,row in df[df["CRESCIMENTO"]>0].nlargest(8,"CRESCIMENTO").iterrows():
            st.markdown(f"""<div class="cli-card positivo"><div>
                <div class="cli-name">{row["CLIENTE_FINAL"]}</div>
                <div class="cli-sub">{row.get("CIDADE","")} · Cresc: {_p(row["CRESCIMENTO"])} · {_m(row["REALIZADO"])}</div>
            </div><span class="badge badge-green">+{_p(row["CRESCIMENTO"])}</span></div>""",unsafe_allow_html=True)

    # Roteiro
    st.markdown("<div class='sec-title'>🚀 Roteiro do dia</div>",unsafe_allow_html=True)
    rot=pd.concat([df[df["PRIORIDADE"]=="🔴 URGENTE"].head(3),df[df["PRIORIDADE"]=="🟡 ATENÇÃO"].head(2)])
    if not rot.empty:
        cr=st.columns(min(len(rot),5))
        for i,(_,row) in enumerate(rot.iterrows()):
            with cr[i]:
                cor="#FEF2F2" if "URGENTE" in row["PRIORIDADE"] else "#FFFBEB"
                brd="#C62828" if "URGENTE" in row["PRIORIDADE"] else "#F9A825"
                st.markdown(f"""<div style="background:{cor};border:1px solid {brd};border-radius:12px;
                    padding:14px;text-align:center;">
                    <div style="font-size:11px;font-weight:700;color:#64748B;">#{i+1}</div>
                    <div style="font-size:13px;font-weight:800;color:#0D1B2A;margin:4px 0;">{row["CLIENTE_FINAL"][:20]}</div>
                    <div style="font-size:11px;color:#64748B;">{row.get("CIDADE","")}</div>
                    <div style="font-size:14px;font-weight:700;color:{brd};margin-top:6px;">{_m(row["GAP"])}</div>
                    <div style="font-size:10px;color:#94A3B8;margin-top:3px;">{row["ACAO"]}</div>
                </div>""",unsafe_allow_html=True)

    # Resumo por cidade
    st.markdown("<div class='sec-title'>📍 Resumo por cidade</div>",unsafe_allow_html=True)
    res=df.groupby("CIDADE",dropna=False).agg(
        CLIENTES=("CLIENTE_FINAL","count"),
        URGENTES=("PRIORIDADE",lambda x:(x=="🔴 URGENTE").sum()),
        ATENCAO =("PRIORIDADE",lambda x:(x=="🟡 ATENÇÃO").sum()),
        META=("META","sum"),REALIZADO=("REALIZADO","sum"),
        PROJECAO=("PROJECAO","sum"),GAP=("GAP","sum"),
    ).reset_index().sort_values(["URGENTES","GAP"],ascending=[False,True])
    st.dataframe(res.style.format({"META":_m,"REALIZADO":_m,"PROJECAO":_m,"GAP":_m})
        .applymap(_cn,subset=["GAP"]),use_container_width=True,hide_index=True)

# ══════════════════════════════════════════════════════════════
# CLIENTE A CLIENTE
# ══════════════════════════════════════════════════════════════
elif PAG=="cliente":
    st.markdown('<div class="app-header"><h1>👤 Cliente a Cliente</h1><p>Análise individual · Histórico · Plano de ataque automático</p></div>',unsafe_allow_html=True)
    _bloco_upload()
    if "df" not in st.session_state: _sem_dados(); st.stop()
    df=st.session_state["df"]

    cb,cs=st.columns([1,2])
    with cb: busca=st.text_input("🔍 Buscar",key="cc_b")
    clis=df["CLIENTE_FINAL"].astype(str).unique().tolist()
    if busca: clis=[c for c in clis if busca.upper() in c.upper()]
    if not clis: st.warning("Nenhum cliente."); st.stop()
    with cs: cnome=st.selectbox("Cliente",clis,key="cc_s")

    d=df[df["CLIENTE_FINAL"].astype(str)==cnome].iloc[0]
    cb2="#C62828" if "URGENTE" in str(d["PRIORIDADE"]) else "#F9A825" if "ATENÇÃO" in str(d["PRIORIDADE"]) else "#1B8A3D"

    st.markdown(f"""<div style="background:white;border:1px solid #E5E7EB;border-left:5px solid {cb2};
        border-radius:14px;padding:18px 22px;margin-bottom:16px;
        display:flex;justify-content:space-between;align-items:center;">
        <div>
            <div style="font-size:20px;font-weight:800;color:#0D1B2A;">{d["CLIENTE_FINAL"]}</div>
            <div style="font-size:13px;color:#64748B;margin-top:3px;">
                📍 {d.get("CIDADE","")} &nbsp;|&nbsp; {d.get("ENDERECO","")} &nbsp;|&nbsp; {d["BANDEIRA_FINAL"]}
            </div>
        </div>
        <div style="text-align:right;">
            <div style="font-size:11px;font-weight:700;color:#64748B;text-transform:uppercase;">Prioridade</div>
            <div style="font-size:16px;font-weight:800;color:{cb2};">{d["PRIORIDADE"]}</div>
        </div>
    </div>""",unsafe_allow_html=True)

    k1,k2,k3,k4,k5=st.columns(5)
    k1.metric("Meta",_m(d["META"])); k2.metric("Realizado",_m(d["REALIZADO"]))
    k3.metric("Projeção",_m(d["PROJECAO"])); k4.metric("GAP",_m(d["GAP"]))
    k5.metric("% GAP",_p(d["PERC_GAP"]))

    c1,c2=st.columns(2)
    with c1: st.metric("Crescimento vs 2025",_p(d["CRESCIMENTO"])); st.metric("Valor 2025",_m(d.get("VALOR_2025",0)))
    with c2: st.metric("Cidade",d.get("CIDADE","—")); st.metric("Supervisor",d.get("SUPERVISOR_CMK","—"))

    # Histórico
    st.markdown("<div class='sec-title'>📈 Histórico de importações</div>",unsafe_allow_html=True)
    hist=_ler_hist_cliente(cnome)
    if hist.empty: st.info("Histórico disponível após múltiplas importações.")
    elif HAS_PLOTLY and len(hist)>1:
        fig=go.Figure()
        fig.add_scatter(x=hist["importado_em"],y=hist["realizado"],name="Realizado",
            mode="lines+markers",line=dict(color="#1B8A3D",width=2),marker=dict(size=6))
        fig.add_scatter(x=hist["importado_em"],y=hist["meta"],name="Meta",
            mode="lines",line=dict(color="#0D1B2A",width=1.5,dash="dash"))
        fig.add_scatter(x=hist["importado_em"],y=hist["projecao"],name="Projeção",
            mode="lines+markers",line=dict(color="#1D4ED8",width=2),marker=dict(size=5))
        fig.update_layout(height=250,margin=dict(l=0,r=0,t=8,b=0),
            plot_bgcolor="white",paper_bgcolor="white",
            legend=dict(orientation="h",y=-0.25),
            yaxis=dict(tickprefix="R$ ",tickformat=",.0f",gridcolor="#F1F5F9"))
        st.plotly_chart(fig,use_container_width=True)
    else:
        st.dataframe(hist[["importado_em","meta","realizado","projecao","gap","prioridade"]],
                     use_container_width=True,hide_index=True)

    # Plano de ataque
    st.markdown("<div class='sec-title'>🎯 Plano de ataque automático</div>",unsafe_allow_html=True)
    pr=str(d["PRIORIDADE"])
    if "URGENTE" in pr:
        nivel="🔴 URGENTE"; cor_n="#C62828"
        intro=f"Situação crítica. GAP de **{_m(d['GAP'])}** ({_p(d['PERC_GAP'])}). Visita imediata."
        acoes=["1️⃣ Visitar pessoalmente — prioridade máxima",
               "2️⃣ Apresentar produtos com maior queda",
               "3️⃣ Verificar débitos antes da visita",
               "4️⃣ Levar proposta com condições especiais",
               "5️⃣ Identificar se compra do concorrente"]
        ab=f'"Vim porque temos GAP de {_m(d["GAP"])} este mês. Quero entender e resolver juntos."'
    elif "ATENÇÃO" in pr:
        nivel="🟡 ATENÇÃO"; cor_n="#F9A825"
        intro=f"Cliente em atenção. GAP de **{_m(d['GAP'])}**. Ainda dá para recuperar."
        acoes=["1️⃣ Agendar visita até fim da semana",
               "2️⃣ Verificar campanhas ativas",
               "3️⃣ Apresentar mix de maior margem",
               "4️⃣ Checar produtos que parou de comprar",
               "5️⃣ Reforçar positivação de famílias"]
        ab=f'"Meta {_m(d["META"])}, realizou {_m(d["REALIZADO"])} — vamos acelerar juntos."'
    else:
        nivel="🟢 POSITIVO"; cor_n="#1B8A3D"
        intro=f"Cliente em crescimento! Realizado **{_m(d['REALIZADO'])}**. Foco em ampliar ticket."
        acoes=["1️⃣ Manter visita de manutenção",
               "2️⃣ Apresentar produtos novos",
               "3️⃣ Explorar mix faltante",
               "4️⃣ Verificar laboratórios extras",
               "5️⃣ Trabalhar famílias ausentes"]
        ab=f'"Crescimento de {_p(d["CRESCIMENTO"])}! Trouxe novidades para continuar crescendo."'

    st.markdown(f"""<div class="section-card">
        <div style="font-size:14px;font-weight:800;color:{cor_n};margin-bottom:10px;">NÍVEL: {nivel}</div>
        <p style="font-size:13px;color:#334155;line-height:1.7;">{intro}</p>
        <div style="background:#F8FAFC;border-radius:10px;padding:14px;margin-top:10px;">
            <div style="font-size:12px;font-weight:700;color:#64748B;margin-bottom:8px;text-transform:uppercase;">Ações recomendadas</div>
            {''.join(f'<div style="font-size:13px;color:#0D1B2A;padding:4px 0;">{a}</div>' for a in acoes)}
        </div>
        <div style="background:#F0FDF4;border:1px solid #86EFAC;border-radius:10px;padding:14px;margin-top:10px;">
            <div style="font-size:12px;font-weight:700;color:#166534;margin-bottom:6px;">💬 ABORDAGEM SUGERIDA</div>
            <div style="font-size:13px;color:#15803D;font-style:italic;line-height:1.6;">{ab}</div>
        </div>
    </div>""",unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# COTABOT
# ══════════════════════════════════════════════════════════════
elif PAG=="cotabot":
    st.markdown('<div class="app-header"><h1>💰 CotaBot</h1><p>Cruzamento por EAN · Preenchimento automático · Download pronto</p></div>',unsafe_allow_html=True)

    def _escrever_xlsx(f,aba,hrow,pcol,precos):
        f.seek(0); wb=load_workbook(f); ws=wb[aba]
        for i,pv in enumerate(precos):
            ws.cell(row=hrow+2+i,column=pcol+1,value=float(pv) if pd.notna(pv) else None)
        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

    st.markdown("#### 1. Base da empresa")
    base_f=st.file_uploader("Base de produtos (EAN + preço + estoque)",type=["xlsx","xls"],key="c_base")
    base_df=None
    if base_f:
        braw=_xl(base_f); bh=_det_h(braw,["codigo ean","código ean","descricao","laboratorio","st","preco nf","estoque"])
        base_df=_mk(braw,bh); st.dataframe(base_df.head(8),use_container_width=True)

    st.markdown("---"); st.markdown("#### 2. Cotação do cliente")
    cot_f=st.file_uploader("Planilha de cotação",type=["xlsx","xls"],key="c_cot")
    cot_df=None; craw=None; aba=None; hrow=None
    if cot_f:
        if cot_f.name.lower().endswith(".xlsx"):
            cot_f.seek(0); wb2=load_workbook(cot_f,read_only=True); abas=wb2.sheetnames; wb2.close()
        else: abas=[0]
        aba=st.selectbox("Aba",abas,key="c_aba") if len(abas)>1 else abas[0]
        craw=_xl(cot_f,sheet_name=aba)
        ch=_det_h(craw,["ean","codigo ean","produto","descricao","qtd","preco","fabricante"])
        hrow=st.number_input("Linha do cabeçalho",min_value=1,max_value=max(1,len(craw)),value=int(ch+1),step=1,key="c_h")-1
        cot_df=_mk(craw,hrow); st.dataframe(cot_df.head(8),use_container_width=True)

    if base_df is not None and cot_df is not None:
        st.markdown("---"); st.subheader("3. Associar colunas")
        ob=["-- Selecionar --"]+list(base_df.columns); oc=["-- Selecionar --"]+list(cot_df.columns)
        be=_fc(base_df.columns,["ean","codigo ean","gtin"]); bp=_fc(base_df.columns,["preço real","preco real","preço final"])
        bs=_fc(base_df.columns,["st","valor st"]); bn=_fc(base_df.columns,["preço nf","preco nf"])
        bst=_fc(base_df.columns,["estoque","saldo","disponivel"])
        ce=_fc(cot_df.columns,["ean","codigo ean","gtin"]); cp2=_fc(cot_df.columns,["preço","preco","valor","price"])
        s1,s2=st.columns(2)
        with s1:
            st.markdown("**Base**")
            col_be=st.selectbox("EAN base",ob,index=ob.index(be) if be in ob else 0,key="c_be")
            modo=st.radio("Preço",["Usar PREÇO REAL","ST + PREÇO NF"],index=0 if bp else 1,key="c_modo")
            if modo=="Usar PREÇO REAL":
                col_bp=st.selectbox("PREÇO REAL",ob,index=ob.index(bp) if bp in ob else 0,key="c_bp"); col_bs=col_bn=None
            else:
                col_bs=st.selectbox("ST",ob,index=ob.index(bs) if bs in ob else 0,key="c_bs")
                col_bn=st.selectbox("PREÇO NF",ob,index=ob.index(bn) if bn in ob else 0,key="c_bn"); col_bp=None
            est_min=st.number_input("Estoque mínimo",min_value=0,value=1,step=1,key="c_est"); col_bst=bst
        with s2:
            st.markdown("**Cotação**")
            col_ce=st.selectbox("EAN cotação",oc,index=oc.index(ce) if ce in oc else 0,key="c_ce")
            col_cp=st.selectbox("PREÇO cotação",oc,index=oc.index(cp2) if cp2 in oc else 0,key="c_cp")
        st.markdown("---")
        if st.button("⚡ Processar cotação",use_container_width=True,key="c_proc"):
            try:
                bp3=base_df.copy(); cp3=cot_df.copy()
                bp3[col_be]=_leia(bp3[col_be]); cp3[col_ce]=_leia(cp3[col_ce])
                if col_bst: bp3[col_bst]=pd.to_numeric(bp3[col_bst],errors="coerce").fillna(0)
                if modo=="Usar PREÇO REAL": bp3["_P_"]=_conv(bp3[col_bp])
                else:
                    bp3["_S_"]=_conv(bp3[col_bs]); bp3["_N_"]=_conv(bp3[col_bn])
                    bp3["_P_"]=bp3["_S_"].fillna(0)+bp3["_N_"].fillna(0)
                bf=bp3[bp3[col_bst]>=est_min].copy() if col_bst else bp3.copy()
                bm=bf[[col_be,"_P_"]].drop_duplicates(subset=[col_be])
                res=cp3.merge(bm,left_on=col_ce,right_on=col_be,how="left")
                nums=res["_P_"].tolist(); prev=[_fmt(x) for x in nums]
                tot=len(cp3); enc=int(pd.notna(res["_P_"]).sum())
                a,b,c=st.columns(3); a.metric("Total",tot); b.metric("✅ Encontrados",enc); c.metric("❌ Não encontrados",tot-enc)
                pv=cp3.copy(); pv[col_cp]=pv[col_cp].astype("object"); pv[col_cp]=prev
                st.dataframe(pv.head(30),use_container_width=True)
                cab_row=craw.iloc[hrow].fillna("").astype(str).tolist()
                pidx=next((i for i,v in enumerate(cab_row) if str(v).strip()==str(col_cp).strip()),None)
                if pidx is None: st.error("Coluna de preço não localizada."); st.stop()
                if cot_f.name.lower().endswith(".xlsx"):
                    out=_escrever_xlsx(cot_f,aba,hrow,pidx,nums)
                    st.download_button("⬇️ Baixar cotação preenchida",out,"cotacao_preenchida.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,key="c_dl")
                else:
                    rb=craw.copy(); rb[pidx]=rb[pidx].astype("object")
                    for i,pv2 in enumerate(nums): rb.iat[hrow+1+i,pidx]=float(pv2) if pd.notna(pv2) else None
                    buf=io.BytesIO()
                    with pd.ExcelWriter(buf,engine="openpyxl") as w: rb.to_excel(w,index=False,header=False)
                    buf.seek(0)
                    st.download_button("⬇️ Baixar",buf.getvalue(),"cotacao_preenchida.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,key="c_dl2")
            except Exception as e: st.error(f"Erro: {e}")
    else: st.info("Suba a base e a cotação para configurar.")

# ══════════════════════════════════════════════════════════════
# COBRANÇA
# ══════════════════════════════════════════════════════════════
elif PAG=="cobranca":
    st.markdown('<div class="app-header"><h1>🧾 Cobrança Inteligente</h1><p>Clientes vencidos · Dias em atraso · Prioridade de cobrança</p></div>',unsafe_allow_html=True)
    cob_f=st.file_uploader("Planilha de cobrança (.xlsx)",type=["xlsx","xls"],key="cob_file")
    if cob_f is None:
        st.info("📨 Suba sua planilha de cobrança para ativar este módulo.")
        st.markdown("**Este módulo vai mostrar:**\n- ✅ Clientes vencidos\n- ✅ Dias em atraso e valor\n- ✅ Prioridade automática\n- ✅ Sugestão de abordagem")
    else:
        try: df_c=pd.read_excel(cob_f)
        except: st.error("Erro ao carregar."); st.stop()
        df_c.columns=[_ln(c) for c in df_c.columns]
        st.dataframe(df_c.head(6),use_container_width=True)
        opcoes=["-- Selecionar --"]+list(df_c.columns)
        col_cli=_ac(df_c,["CLIENTE","RAZAO","NOME"]); col_val=_ac(df_c,["VALOR","TOTAL","SALDO","PENDENTE"])
        col_dias=_ac(df_c,["DIAS","DIAS ATRASO","ATRASO"]); col_cid=_ac(df_c,["CIDADE","MUNICIPIO"])
        c1,c2,c3=st.columns(3)
        with c1:
            col_cli_s=st.selectbox("Cliente",opcoes,index=opcoes.index(col_cli) if col_cli in opcoes else 0,key="cob_cli")
            col_val_s=st.selectbox("Valor",  opcoes,index=opcoes.index(col_val)  if col_val  in opcoes else 0,key="cob_val")
        with c2:
            col_dia_s=st.selectbox("Dias atraso",opcoes,index=opcoes.index(col_dias) if col_dias in opcoes else 0,key="cob_dia")
            col_cid_s=st.selectbox("Cidade",     opcoes,index=opcoes.index(col_cid)  if col_cid  in opcoes else 0,key="cob_cid")
        if st.button("⚡ Processar cobrança",use_container_width=True,key="cob_proc"):
            try:
                df2=df_c.copy()
                df2["CLI"]=df2[col_cli_s].astype(str).str.strip() if col_cli_s!="-- Selecionar --" else "SEM NOME"
                df2["VAL"]=pd.to_numeric(df2[col_val_s],errors="coerce").fillna(0) if col_val_s!="-- Selecionar --" else 0.0
                df2["DIA"]=pd.to_numeric(df2[col_dia_s],errors="coerce").fillna(0) if col_dia_s!="-- Selecionar --" else 0.0
                df2["CID"]=df2[col_cid_s].astype(str) if col_cid_s!="-- Selecionar --" else ""
                def pc(row):
                    if row["DIA"]>=60: return "🔴 CRÍTICO"
                    if row["DIA"]>=30: return "🟡 ATENÇÃO"
                    return "🟢 RECENTE"
                df2["PRIO"]=df2.apply(pc,axis=1); df2=df2.sort_values(["DIA","VAL"],ascending=[False,False])
                k1,k2,k3,k4=st.columns(4)
                with k1: st.markdown(f'<div class="kpi-card red"><div class="kpi-label">Total em aberto</div><div class="kpi-value">{_m(df2["VAL"].sum())}</div></div>',unsafe_allow_html=True)
                with k2: st.markdown(f'<div class="kpi-card red"><div class="kpi-label">Críticos +60d</div><div class="kpi-value">{int((df2["PRIO"]=="🔴 CRÍTICO").sum())}</div></div>',unsafe_allow_html=True)
                with k3: st.markdown(f'<div class="kpi-card amber"><div class="kpi-label">Atenção 30-60d</div><div class="kpi-value">{int((df2["PRIO"]=="🟡 ATENÇÃO").sum())}</div></div>',unsafe_allow_html=True)
                with k4: st.markdown(f'<div class="kpi-card navy"><div class="kpi-label">Total clientes</div><div class="kpi-value">{len(df2)}</div></div>',unsafe_allow_html=True)
                def cpc(v):
                    if "CRÍTICO" in str(v): return "background-color:#F8D7DA;color:#842029;font-weight:bold"
                    if "ATENÇÃO" in str(v): return "background-color:#FFF3CD;color:#664D03;font-weight:bold"
                    return "background-color:#D1E7DD;color:#0F5132;font-weight:bold"
                ex=df2[["CLI","CID","PRIO","VAL","DIA"]].copy()
                ex.columns=["CLIENTE","CIDADE","PRIORIDADE","VALOR","DIAS"]
                st.dataframe(ex.style.format({"VALOR":_m,"DIAS":"{:.0f}"}).applymap(cpc,subset=["PRIORIDADE"]),
                    use_container_width=True,hide_index=True,height=480)
                st.markdown("<div class='sec-title'>🔴 Críticos</div>",unsafe_allow_html=True)
                for _,row in df2[df2["PRIO"]=="🔴 CRÍTICO"].head(5).iterrows():
                    st.markdown(f"""<div class="cli-card urgente"><div>
                        <div class="cli-name">{row["CLI"]}</div>
                        <div class="cli-sub">{row.get("CID","")} · {int(row["DIA"])} dias · {_m(row["VAL"])}</div>
                    </div><span class="badge badge-red">🔴 {int(row["DIA"])}d</span></div>""",unsafe_allow_html=True)
            except Exception as e: st.error(f"Erro: {e}")

# ══════════════════════════════════════════════════════════════
# PAINEL DE VISITAS
# ══════════════════════════════════════════════════════════════
elif PAG=="visitas":
    st.markdown('<div class="app-header"><h1>📍 Painel de Visitas</h1><p>GPS estratégico · Ranking por cidade · Rota inteligente</p></div>',unsafe_allow_html=True)
    _bloco_upload()
    if "df" not in st.session_state: _sem_dados(); st.stop()
    df=st.session_state["df"].copy()

    f1,f2,f3=st.columns(3)
    with f1:
        cids=["Todas"]+sorted(df["CIDADE"].dropna().astype(str).unique().tolist())
        f_cid=st.selectbox("📍 Município",cids,key="v_cid")
    with f2: f_prio=st.selectbox("🎯 Prioridade",["Todas","🔴 URGENTE","🟡 ATENÇÃO","🟢 AUMENTAR VENDA"],key="v_prio")
    with f3: busca=st.text_input("🔍 Buscar cliente",key="v_b")

    if f_cid!="Todas": df=df[df["CIDADE"].astype(str)==f_cid]
    if f_prio!="Todas": df=df[df["PRIORIDADE"]==f_prio]
    if busca: df=df[df["CLIENTE_FINAL"].str.contains(busca,case=False,na=False)]
    if df.empty: st.warning("Nenhum cliente com os filtros."); st.stop()

    urg=int((df["PRIORIDADE"]=="🔴 URGENTE").sum()); ate=int((df["PRIORIDADE"]=="🟡 ATENÇÃO").sum())
    k1,k2,k3,k4=st.columns(4)
    with k1: st.markdown(f'<div class="kpi-card navy"><div class="kpi-label">Clientes</div><div class="kpi-value">{len(df)}</div></div>',unsafe_allow_html=True)
    with k2: st.markdown(f'<div class="kpi-card red"><div class="kpi-label">Urgentes</div><div class="kpi-value">{urg}</div><div class="kpi-sub">{ate} em atenção</div></div>',unsafe_allow_html=True)
    cg2="red" if df["GAP"].sum()<0 else "green"
    with k3: st.markdown(f'<div class="kpi-card {cg2}"><div class="kpi-label">GAP</div><div class="kpi-value">{_m(df["GAP"].sum())}</div></div>',unsafe_allow_html=True)
    with k4: st.markdown(f'<div class="kpi-card blue"><div class="kpi-label">Meta</div><div class="kpi-value">{_m(df["META"].sum())}</div></div>',unsafe_allow_html=True)
    st.markdown("<div style='margin-top:10px;'></div>",unsafe_allow_html=True)

    if f_cid=="Todas" and HAS_PLOTLY:
        st.markdown("<div class='sec-title'>🏙️ Cidades por urgência</div>",unsafe_allow_html=True)
        rc=df.groupby("CIDADE",dropna=False).agg(
            URGENTES=("PRIORIDADE",lambda x:(x=="🔴 URGENTE").sum()),
            ATENCAO =("PRIORIDADE",lambda x:(x=="🟡 ATENÇÃO").sum()),
        ).reset_index().sort_values("URGENTES",ascending=False).head(10)
        fig=go.Figure()
        fig.add_bar(name="Urgentes",x=rc["CIDADE"],y=rc["URGENTES"],marker_color="#C62828")
        fig.add_bar(name="Atenção", x=rc["CIDADE"],y=rc["ATENCAO"], marker_color="#F9A825")
        fig.update_layout(height=240,margin=dict(l=0,r=0,t=8,b=60),
            plot_bgcolor="white",paper_bgcolor="white",barmode="stack",
            showlegend=True,legend=dict(orientation="h",y=-0.4),
            xaxis=dict(tickangle=-30,tickfont=dict(size=11)),yaxis=dict(gridcolor="#F1F5F9"))
        st.plotly_chart(fig,use_container_width=True)

    st.markdown("<div class='sec-title'>🎯 Clientes prioritários</div>",unsafe_allow_html=True)
    tb=df[["CLIENTE_FINAL","CIDADE","PRIORIDADE","META","REALIZADO","GAP","PERC_GAP","CRESCIMENTO","ENDERECO","ACAO"]].copy()
    tb.columns=["CLIENTE","CIDADE","PRIORIDADE","META R$","REALIZADO R$","GAP R$","% GAP","CRESC.","ENDEREÇO","AÇÃO"]
    st.dataframe(tb.style.format({"META R$":_m,"REALIZADO R$":_m,"GAP R$":_m,"% GAP":_p,"CRESC.":_p})
        .applymap(_cp,subset=["PRIORIDADE"]).applymap(_cn,subset=["GAP R$","CRESC."]),
        use_container_width=True,hide_index=True,height=480)

    st.markdown("<div class='sec-title'>🗺️ Roteiro do dia</div>",unsafe_allow_html=True)
    n=st.slider("Visitas",3,20,10,key="v_n")
    if st.button("🚀 Gerar roteiro",use_container_width=True,key="v_rot"):
        rot=pd.concat([df[df["PRIORIDADE"]=="🔴 URGENTE"].sort_values(["CIDADE","GAP"]),
                       df[df["PRIORIDADE"]=="🟡 ATENÇÃO"].sort_values(["CIDADE","GAP"]),
                       df[df["PRIORIDADE"]=="🟢 AUMENTAR VENDA"].sort_values("GAP",ascending=False)]).head(n)
        if rot.empty: st.warning("Nenhum cliente.")
        else:
            st.success(f"Roteiro com {len(rot)} visitas!")
            cid_atual=None
            for i,(_,row) in enumerate(rot.iterrows(),1):
                if row.get("CIDADE")!=cid_atual:
                    cid_atual=row.get("CIDADE",""); st.markdown(f"#### 📍 {cid_atual}")
                cor="#FEF2F2" if "URGENTE" in row["PRIORIDADE"] else "#FFFBEB" if "ATENÇÃO" in row["PRIORIDADE"] else "#F0FDF4"
                brd="#C62828" if "URGENTE" in row["PRIORIDADE"] else "#F9A825" if "ATENÇÃO" in row["PRIORIDADE"] else "#1B8A3D"
                st.markdown(f"""<div style="background:{cor};border-left:4px solid {brd};border-radius:8px;
                    padding:10px 14px;margin-bottom:6px;">
                    <strong>#{i} {row["CLIENTE_FINAL"]}</strong> · {row["PRIORIDADE"]} ·
                    GAP: <strong>{_m(row["GAP"])}</strong> · {row.get("ENDERECO","")}
                    <span style="float:right;font-size:12px;color:#64748B;">{row["ACAO"]}</span>
                </div>""",unsafe_allow_html=True)
