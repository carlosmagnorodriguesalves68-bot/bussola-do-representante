import streamlit as st
import pandas as pd
import unicodedata
import re
import io
from openpyxl import load_workbook

st.set_page_config(
    page_title="Bússola do Representante",
    page_icon="🧭",
    layout="wide",
)

# ══════════════════════════════════════════════════════════════
# CSS GLOBAL
# ══════════════════════════════════════════════════════════════
st.markdown("""
<style>
section[data-testid="stSidebar"] { display: none !important; }
header[data-testid="stHeader"]   { display: none !important; }
.block-container { padding-top: 0 !important; padding-bottom: 2rem; max-width: 1400px; }
.navbar {
    background: linear-gradient(90deg, #0D1B2A 0%, #13293D 70%, #0D1B2A 100%);
    padding: 0 28px; display: flex; align-items: center;
    border-bottom: 3px solid #C62828; height: 56px; margin-bottom: 0;
}
.navbar-brand { font-size: 17px; font-weight: 800; color: white; margin-right: 32px; white-space: nowrap; }
.navbar-brand span { color: #C62828; }
.kpi-card { background: white; border: 1px solid #E5E7EB; border-radius: 14px; padding: 16px 18px; height: 100%; }
.kpi-card.red   { border-left: 4px solid #C62828; }
.kpi-card.green { border-left: 4px solid #1B8A3D; }
.kpi-card.amber { border-left: 4px solid #F9A825; }
.kpi-card.blue  { border-left: 4px solid #1D4ED8; }
.kpi-card.navy  { border-left: 4px solid #0D1B2A; }
.kpi-label { font-size: 11px; color: #64748B; font-weight: 700; text-transform: uppercase; letter-spacing: 0.6px; margin-bottom: 6px; }
.kpi-value { font-size: 24px; font-weight: 800; color: #0D1B2A; line-height: 1.1; }
.kpi-sub   { font-size: 11px; color: #94A3B8; margin-top: 3px; }
.cli-card { background: white; border: 1px solid #E5E7EB; border-radius: 12px; padding: 12px 14px; margin-bottom: 8px; display: flex; justify-content: space-between; align-items: center; }
.cli-card.urgente { border-left: 4px solid #C62828; }
.cli-card.atencao { border-left: 4px solid #F9A825; }
.cli-name { font-size: 13px; font-weight: 700; color: #0D1B2A; }
.cli-sub  { font-size: 11px; color: #64748B; margin-top: 2px; }
.badge { font-size: 11px; font-weight: 700; padding: 3px 10px; border-radius: 20px; white-space: nowrap; }
.badge-red   { background: #FEF2F2; color: #C62828; }
.badge-amber { background: #FFFBEB; color: #B45309; }
.badge-green { background: #F0FDF4; color: #1B8A3D; }
.sec-title { font-size: 14px; font-weight: 700; color: #0D1B2A; margin: 20px 0 10px; padding-bottom: 6px; border-bottom: 2px solid #E5E7EB; }
.section-card { background: white; border: 1px solid #E5E7EB; border-radius: 16px; padding: 18px; margin-bottom: 18px; }
.app-header { background: linear-gradient(90deg,#0D1B2A 0%,#13293D 60%,#C62828 100%); padding: 20px 28px; border-radius: 16px; margin-bottom: 20px; color: white; }
.app-header h1 { color: white; margin-bottom: 4px; font-size: 26px; }
.app-header p  { color: #E5E7EB; font-size: 13px; margin: 0; }
div[data-testid="stMetricLabel"] { color: #475569; font-weight: 600; }
div[data-testid="stMetricValue"] { color: #0D1B2A; font-weight: 800; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# ESTADO + NAVBAR
# ══════════════════════════════════════════════════════════════
if "pagina" not in st.session_state:
    st.session_state["pagina"] = "dashboard"

st.markdown('<div class="navbar"><div class="navbar-brand">🧭 Bússola do <span>Representante</span></div></div>', unsafe_allow_html=True)

nc1,nc2,nc3,nc4,nc5,nc6 = st.columns([2.5,1.4,1.4,1.4,1.4,1.4])
with nc2:
    if st.button("🧭 Dashboard", use_container_width=True,
                 type="primary" if st.session_state["pagina"]=="dashboard" else "secondary"):
        st.session_state["pagina"]="dashboard"; st.rerun()
with nc3:
    if st.button("📡 Radar", use_container_width=True,
                 type="primary" if st.session_state["pagina"]=="radar" else "secondary"):
        st.session_state["pagina"]="radar"; st.rerun()
with nc4:
    if st.button("💰 CotaBot", use_container_width=True,
                 type="primary" if st.session_state["pagina"]=="cotabot" else "secondary"):
        st.session_state["pagina"]="cotabot"; st.rerun()
with nc5:
    if st.button("🧾 Cobrança", use_container_width=True,
                 type="primary" if st.session_state["pagina"]=="cobranca" else "secondary"):
        st.session_state["pagina"]="cobranca"; st.rerun()
with nc6:
    if st.button("📍 Visitas", use_container_width=True,
                 type="primary" if st.session_state["pagina"]=="visitas" else "secondary"):
        st.session_state["pagina"]="visitas"; st.rerun()

st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
pagina = st.session_state["pagina"]

# ══════════════════════════════════════════════════════════════
# FUNÇÕES GLOBAIS
# ══════════════════════════════════════════════════════════════
def _ln(txt):
    txt = str(txt).strip().upper()
    return unicodedata.normalize("NFKD", txt).encode("ASCII","ignore").decode("utf-8")

def _num(v):
    if pd.isna(v): return 0.0
    if isinstance(v,(int,float)): return float(v)
    v=str(v).strip().replace("R$","").replace("%","").replace(".","").replace(",",".")
    try: return float(v)
    except: return 0.0

def _m(v):
    try: return f"R$ {float(v):,.0f}".replace(",",".")
    except: return "R$ 0"

def _p(v):
    try: return f"{float(v)*100:,.1f}%".replace(",","X").replace(".",",").replace("X",".")
    except: return "0,0%"

def _p2(v):
    try: return f"{float(v)*100:,.2f}%".replace(",","X").replace(".",",").replace("X",".")
    except: return "0,00%"

def _ac(df, ops, exata=False):
    for o in ops:
        for c in df.columns:
            if _ln(c)==_ln(o): return c
    if not exata:
        for o in ops:
            for c in df.columns:
                if _ln(o) in _ln(c): return c
    return None

def _prio(row):
    if row["GAP"]>=0: return "🟢 AUMENTAR VENDA"
    if row["PERC_GAP"]<=-0.20: return "🔴 URGENTE"
    return "🟡 ATENÇÃO"

def _acao(row):
    if row["GAP"]>=0: return "Aumentar venda"
    if row["PERC_GAP"]<=-0.20: return "Recuperar urgente"
    return "Acompanhar"

def _cp(v):
    if "URGENTE"  in str(v): return "background-color:#F8D7DA;color:#842029;font-weight:bold"
    if "ATENÇÃO"  in str(v): return "background-color:#FFF3CD;color:#664D03;font-weight:bold"
    if "AUMENTAR" in str(v): return "background-color:#D1E7DD;color:#0F5132;font-weight:bold"
    return ""

def _cn(v):
    try:
        x=float(v)
        if x<0: return "color:#C62828;font-weight:bold"
        if x>0: return "color:#1B8A3D;font-weight:bold"
    except: pass
    return ""

def carregar_curva(curva_file, cmk_file):
    try:    curva=pd.read_excel(curva_file, sheet_name="DADOS")
    except: curva=pd.read_excel(curva_file)
    cmk=pd.read_excel(cmk_file)
    curva.columns=[_ln(c) for c in curva.columns]
    cmk.columns  =[_ln(c) for c in cmk.columns]

    col_cliente=_ac(curva,["CLIENTE"],exata=True)
    col_cnpj_c =_ac(curva,["CNPJ"],exata=True)
    col_cnpj_m =_ac(cmk,  ["CNPJ"],exata=True)
    col_bandeira=_ac(curva,["BANDEIRA"],exata=True)
    col_sup    =_ac(curva,["SUPERVISOR"])
    col_setor  =_ac(curva,["COD.SETOR","COD SETOR","SETOR"])
    col_meta   =_ac(curva,["META"],exata=True)
    col_real   =_ac(curva,["REAL"],exata=True)
    col_proj   =_ac(curva,["REAL PROJ","REAL PROJ AC"])
    col_gap    =_ac(curva,["DESVIO PROJ"],exata=True)
    col_perc   =_ac(curva,["% DESVIO"],exata=True)
    col_2025   =_ac(curva,["2025"],exata=True)
    col_cresc  =_ac(curva,["% CRESC PROJ","% CRESC PROJ AC","CRESC"])
    col_mpv    =_ac(curva,["META PV"])
    col_pven   =_ac(curva,["P.VEN","P VEN","PVEN"])
    col_dpv    =_ac(curva,["DESVIO PV"])
    col_mcmv   =_ac(curva,["META CMV %","META CMV"])
    col_cmvr   =_ac(curva,["CMV %","CMV REAL","CMV"])
    col_dcmv   =_ac(curva,["DESVIO CMV"])

    df=curva.copy()
    df["META"]       =df[col_meta].apply(_num)   if col_meta   else 0
    df["REALIZADO"]  =df[col_real].apply(_num)   if col_real   else 0
    df["PROJECAO"]   =df[col_proj].apply(_num)   if col_proj   else 0
    df["GAP"]        =df[col_gap].apply(_num)    if col_gap    else df["PROJECAO"]-df["META"]
    df["PERC_GAP"]   =df[col_perc].apply(_num)   if col_perc   else 0
    df["VALOR_2025"] =df[col_2025].apply(_num)   if col_2025   else 0
    df["CRESCIMENTO"]=df[col_cresc].apply(_num)  if col_cresc  else 0
    df["META_PV"]    =df[col_mpv].apply(_num)    if col_mpv    else 0
    df["PRAZO_REAL"] =df[col_pven].apply(_num)   if col_pven   else 0
    df["DESVIO_PRAZO"]=df[col_dpv].apply(_num)   if col_dpv    else 0
    df["META_CMV"]   =df[col_mcmv].apply(_num)   if col_mcmv   else 0
    df["CMV_REAL"]   =df[col_cmvr].apply(_num)   if col_cmvr   else 0
    df["DESVIO_CMV"] =df[col_dcmv].apply(_num)   if col_dcmv   else 0

    df["CLIENTE_FINAL"] =df[col_cliente].astype(str).str.strip() if col_cliente else "SEM NOME"
    df["BANDEIRA_FINAL"]=df[col_bandeira].astype(str).str.strip() if col_bandeira else ""
    df["SETOR_FINAL"]   =df[col_setor].apply(lambda v: (re.search(r"\d+",str(v)) or type("",(),{"group":lambda s,*a:""})()).group()) if col_setor else ""

    if col_cnpj_c and col_cnpj_m:
        df["_CNPJ_"] =df[col_cnpj_c].apply(lambda v: re.sub(r"\D","",str(v)))
        cmk["_CNPJ_"]=cmk[col_cnpj_m].apply(lambda v: re.sub(r"\D","",str(v)))
        col_cid=_ac(cmk,["MUNICIPIO","CIDADE","CITY"])
        col_end=_ac(cmk,["ENDERECO","ENDEREÇO","END","LOGRADOURO"])
        campos={"_CNPJ_":"_CNPJ_"}
        if col_cid: campos[col_cid]="CIDADE"
        if col_end: campos[col_end]="ENDERECO"
        cmk_m=cmk[list(campos.keys())].drop_duplicates("_CNPJ_").rename(columns=campos)
        df=df.merge(cmk_m,on="_CNPJ_",how="left")
    if "CIDADE"   not in df.columns: df["CIDADE"]=""
    if "ENDERECO" not in df.columns: df["ENDERECO"]=""

    df["PRIORIDADE"]=df.apply(_prio,axis=1)
    df["ACAO"]      =df.apply(_acao,axis=1)
    ordem={"🔴 URGENTE":1,"🟡 ATENÇÃO":2,"🟢 AUMENTAR VENDA":3}
    df["ORDEM"]=df["PRIORIDADE"].map(ordem).fillna(9)
    df=df.sort_values(["ORDEM","GAP"],ascending=[True,True])
    return df, col_sup

# ══════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════
if pagina=="dashboard":
    try:
        import plotly.graph_objects as go
        HAS_PLOTLY=True
    except:
        HAS_PLOTLY=False

    with st.expander("📁 Carregar dados do setor", expanded="df_dash" not in st.session_state):
        u1,u2=st.columns(2)
        with u1: curva_f=st.file_uploader("📊 Curva Semanal (.xlsx)",type=["xlsx"],key="d_curva")
        with u2: cmk_f  =st.file_uploader("📍 CMK / Endereço (.xlsx)",type=["xlsx"],key="d_cmk")
        if curva_f and cmk_f:
            df_d,cs_d=carregar_curva(curva_f,cmk_f)
            st.session_state["df_dash"]=df_d
            st.session_state["cs_dash"]=cs_d
            st.success("✅ Dados carregados!")

    if "df_dash" not in st.session_state:
        st.markdown("""
        <div style="background:#F8FAFC;border:2px dashed #CBD5E1;border-radius:16px;
                    padding:60px;text-align:center;margin-top:20px;">
            <div style="font-size:52px;margin-bottom:12px;">🧭</div>
            <div style="font-size:22px;font-weight:800;color:#0D1B2A;margin-bottom:8px;">Bússola do Representante</div>
            <div style="font-size:14px;color:#64748B;">
                Abra o painel acima e suba a <strong>Curva Semanal</strong> + <strong>CMK</strong>
                para ver o dashboard completo do seu setor.
            </div>
        </div>""", unsafe_allow_html=True)
        st.stop()

    df=st.session_state["df_dash"]
    total    =len(df)
    urgentes =len(df[df["PRIORIDADE"]=="🔴 URGENTE"])
    atencao  =len(df[df["PRIORIDADE"]=="🟡 ATENÇÃO"])
    positivos=len(df[df["PRIORIDADE"]=="🟢 AUMENTAR VENDA"])
    meta_t   =df["META"].sum()
    real_t   =df["REALIZADO"].sum()
    proj_t   =df["PROJECAO"].sum()
    gap_t    =df["GAP"].sum()
    pct_meta =(real_t/meta_t*100) if meta_t>0 else 0

    # KPIs
    k1,k2,k3,k4,k5=st.columns(5)
    with k1: st.markdown(f'<div class="kpi-card navy"><div class="kpi-label">Meta total</div><div class="kpi-value">{_m(meta_t)}</div><div class="kpi-sub">{total} clientes</div></div>',unsafe_allow_html=True)
    with k2: st.markdown(f'<div class="kpi-card green"><div class="kpi-label">Realizado</div><div class="kpi-value">{_m(real_t)}</div><div class="kpi-sub">{pct_meta:.1f}% da meta</div></div>',unsafe_allow_html=True)
    with k3: st.markdown(f'<div class="kpi-card blue"><div class="kpi-label">Projeção</div><div class="kpi-value">{_m(proj_t)}</div><div class="kpi-sub">até fim do mês</div></div>',unsafe_allow_html=True)
    cor_g="red" if gap_t<0 else "green"
    with k4: st.markdown(f'<div class="kpi-card {cor_g}"><div class="kpi-label">GAP total</div><div class="kpi-value">{_m(gap_t)}</div><div class="kpi-sub">{"déficit" if gap_t<0 else "superávit"}</div></div>',unsafe_allow_html=True)
    with k5: st.markdown(f'<div class="kpi-card red"><div class="kpi-label">🔴 Urgentes</div><div class="kpi-value">{urgentes}</div><div class="kpi-sub">{atencao} em atenção</div></div>',unsafe_allow_html=True)

    st.markdown("<div style='margin-top:10px;'></div>",unsafe_allow_html=True)

    if HAS_PLOTLY:
        import plotly.graph_objects as go

        g1,g2=st.columns([1.4,1])
        with g1:
            st.markdown("<div class='sec-title'>📊 Faturamento do mês vs meta</div>",unsafe_allow_html=True)
            fig=go.Figure()
            fig.add_bar(name="Meta",      x=["Meta"],      y=[meta_t], marker_color="#0D1B2A")
            fig.add_bar(name="Realizado", x=["Realizado"], y=[real_t], marker_color="#1B8A3D")
            fig.add_bar(name="Projeção",  x=["Projeção"],  y=[proj_t], marker_color="#1D4ED8")
            fig.update_layout(height=250,margin=dict(l=0,r=0,t=8,b=0),
                plot_bgcolor="white",paper_bgcolor="white",showlegend=True,
                legend=dict(orientation="h",y=-0.2),bargap=0.35,
                yaxis=dict(tickprefix="R$ ",tickformat=",.0f",gridcolor="#F1F5F9"))
            st.plotly_chart(fig,use_container_width=True)

        with g2:
            st.markdown("<div class='sec-title'>🎯 Prioridades do setor</div>",unsafe_allow_html=True)
            fig2=go.Figure(go.Pie(
                labels=["🔴 Urgente","🟡 Atenção","🟢 Positivo"],
                values=[urgentes,atencao,positivos],
                marker_colors=["#C62828","#F9A825","#1B8A3D"],
                hole=0.48, textinfo="label+percent", textfont_size=12,
            ))
            fig2.update_layout(height=250,margin=dict(l=0,r=0,t=8,b=0),
                paper_bgcolor="white",showlegend=False)
            st.plotly_chart(fig2,use_container_width=True)

        g3,g4=st.columns([1,1.4])
        with g3:
            st.markdown("<div class='sec-title'>📍 GAP por cidade</div>",unsafe_allow_html=True)
            rc=df.groupby("CIDADE").agg(GAP_TOTAL=("GAP","sum")).reset_index()
            rc=rc[rc["CIDADE"].astype(str).str.strip()!=""].sort_values("GAP_TOTAL").tail(12)
            cores=["#C62828" if v<0 else "#1B8A3D" for v in rc["GAP_TOTAL"]]
            fig3=go.Figure(go.Bar(
                x=rc["GAP_TOTAL"],y=rc["CIDADE"],orientation="h",
                marker_color=cores,
                text=[_m(v) for v in rc["GAP_TOTAL"]],
                textposition="outside",textfont=dict(size=10),
            ))
            fig3.update_layout(height=300,margin=dict(l=0,r=70,t=8,b=0),
                plot_bgcolor="white",paper_bgcolor="white",
                xaxis=dict(tickprefix="R$ ",tickformat=",.0f",gridcolor="#F1F5F9"),
                yaxis=dict(tickfont=dict(size=11)))
            st.plotly_chart(fig3,use_container_width=True)

        with g4:
            st.markdown("<div class='sec-title'>📈 Real vs projeção — top 10 clientes</div>",unsafe_allow_html=True)
            top=df.nlargest(10,"META")[["CLIENTE_FINAL","REALIZADO","PROJECAO"]].copy()
            top["CLIENTE_FINAL"]=top["CLIENTE_FINAL"].str[:18]
            fig4=go.Figure()
            fig4.add_bar(name="Realizado",x=top["CLIENTE_FINAL"],y=top["REALIZADO"],marker_color="#1B8A3D")
            fig4.add_bar(name="Projeção", x=top["CLIENTE_FINAL"],y=top["PROJECAO"], marker_color="#1D4ED8",opacity=0.75)
            fig4.update_layout(height=300,margin=dict(l=0,r=0,t=8,b=70),
                plot_bgcolor="white",paper_bgcolor="white",barmode="group",
                showlegend=True,legend=dict(orientation="h",y=-0.35),
                yaxis=dict(tickprefix="R$ ",tickformat=",.0f",gridcolor="#F1F5F9"),
                xaxis=dict(tickangle=-35,tickfont=dict(size=10)))
            st.plotly_chart(fig4,use_container_width=True)
    else:
        st.warning("Adicione 'plotly' no requirements.txt para ver os gráficos.")

    # Clientes prioritários
    p1,p2=st.columns(2)
    with p1:
        st.markdown("<div class='sec-title'>🔴 Urgentes — ação imediata</div>",unsafe_allow_html=True)
        df_urg=df[df["PRIORIDADE"]=="🔴 URGENTE"].head(8)
        if df_urg.empty: st.success("Nenhum cliente urgente!")
        for _,row in df_urg.iterrows():
            st.markdown(f"""<div class="cli-card urgente">
                <div><div class="cli-name">{row["CLIENTE_FINAL"]}</div>
                <div class="cli-sub">{row.get("CIDADE","")} · GAP: {_m(row["GAP"])} · {_p(row["PERC_GAP"])}</div></div>
                <span class="badge badge-red">Urgente</span></div>""",unsafe_allow_html=True)

    with p2:
        st.markdown("<div class='sec-title'>🟡 Atenção — acompanhar hoje</div>",unsafe_allow_html=True)
        df_ate=df[df["PRIORIDADE"]=="🟡 ATENÇÃO"].head(8)
        if df_ate.empty: st.success("Nenhum em atenção.")
        for _,row in df_ate.iterrows():
            st.markdown(f"""<div class="cli-card atencao">
                <div><div class="cli-name">{row["CLIENTE_FINAL"]}</div>
                <div class="cli-sub">{row.get("CIDADE","")} · GAP: {_m(row["GAP"])} · {_p(row["PERC_GAP"])}</div></div>
                <span class="badge badge-amber">Atenção</span></div>""",unsafe_allow_html=True)

    # Roteiro do dia
    st.markdown("<div class='sec-title'>🚀 Roteiro do dia — top 5 prioridades</div>",unsafe_allow_html=True)
    roteiro=pd.concat([df[df["PRIORIDADE"]=="🔴 URGENTE"].head(3),df[df["PRIORIDADE"]=="🟡 ATENÇÃO"].head(2)])
    if not roteiro.empty:
        cols=st.columns(min(len(roteiro),5))
        for i,(_,row) in enumerate(roteiro.iterrows()):
            with cols[i]:
                cor="#FEF2F2" if "URGENTE" in row["PRIORIDADE"] else "#FFFBEB"
                brd="#C62828" if "URGENTE" in row["PRIORIDADE"] else "#F9A825"
                st.markdown(f"""<div style="background:{cor};border:1px solid {brd};border-radius:12px;
                    padding:14px;text-align:center;">
                    <div style="font-size:11px;font-weight:700;color:#64748B;">#{i+1}</div>
                    <div style="font-size:13px;font-weight:800;color:#0D1B2A;margin:4px 0;">{row["CLIENTE_FINAL"][:20]}</div>
                    <div style="font-size:11px;color:#64748B;">{row.get("CIDADE","")}</div>
                    <div style="font-size:14px;font-weight:700;color:{brd};margin-top:6px;">{_m(row["GAP"])}</div>
                    <div style="font-size:10px;color:#94A3B8;margin-top:3px;">{row["ACAO"]}</div>
                </div>""",unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# RADAR COMERCIAL
# ══════════════════════════════════════════════════════════════
elif pagina=="radar":
    st.markdown('<div class="app-header"><h1>📡 Radar Comercial Inteligente</h1><p>Análise completa — metas, GAP, prioridade e roteiro.</p></div>',unsafe_allow_html=True)
    u1,u2=st.columns(2)
    with u1: curva_f=st.file_uploader("📊 Curva Semanal",type=["xlsx"],key="r_curva")
    with u2: cmk_f  =st.file_uploader("📍 CMK",         type=["xlsx"],key="r_cmk")
    if not curva_f or not cmk_f: st.info("Suba a Curva Semanal e o CMK."); st.stop()

    df,col_sup=carregar_curva(curva_f,cmk_f)
    f1,f2,f3,f4=st.columns(4)
    with f1: busca =st.text_input("🔍 Cliente",key="r_busca")
    with f2:
        cids=["Todas"]+sorted(df["CIDADE"].dropna().astype(str).unique().tolist())
        f_cid=st.selectbox("Cidade",cids,key="r_cid")
    with f3:
        f_prio=st.selectbox("Prioridade",["Todas","🔴 URGENTE","🟡 ATENÇÃO","🟢 AUMENTAR VENDA"],key="r_prio")
    with f4:
        if col_sup:
            sups=["Todos"]+sorted(df[col_sup].dropna().astype(str).unique().tolist())
            f_sup=st.selectbox("Supervisor",sups,key="r_sup")
        else: f_sup="Todos"

    if busca: df=df[df["CLIENTE_FINAL"].str.contains(busca,case=False,na=False)]
    if f_cid!="Todas": df=df[df["CIDADE"].astype(str)==f_cid]
    if f_prio!="Todas": df=df[df["PRIORIDADE"]==f_prio]
    if col_sup and f_sup!="Todos": df=df[df[col_sup].astype(str)==f_sup]
    if df.empty: st.warning("Nenhum cliente."); st.stop()

    c1,c2,c3,c4,c5=st.columns(5)
    c1.metric("Clientes",len(df)); c2.metric("Positivos",len(df[df["GAP"]>=0]))
    c3.metric("Negativos",len(df[df["GAP"]<0])); c4.metric("Urgentes",len(df[df["PRIORIDADE"]=="🔴 URGENTE"]))
    c5.metric("Meta",_m(df["META"].sum()))
    m1,m2,m3,m4=st.columns(4)
    m1.metric("Realizado",_m(df["REALIZADO"].sum())); m2.metric("Projeção",_m(df["PROJECAO"].sum()))
    m3.metric("GAP",_m(df["GAP"].sum()))
    crit=df[df["PRIORIDADE"]=="🔴 URGENTE"]["CIDADE"].value_counts()
    m4.metric("Cidade crítica",crit.idxmax() if not crit.empty else "-")
    st.divider()

    tabela=df[["CLIENTE_FINAL","CIDADE","PRIORIDADE","META","REALIZADO","PROJECAO","GAP","PERC_GAP","VALOR_2025","CRESCIMENTO","ACAO"]].copy()
    tabela.columns=["CLIENTE","CIDADE","PRIORIDADE","META R$","REALIZADO R$","PROJEÇÃO R$","GAP R$","% GAP","2025 R$","CRESC.","AÇÃO"]
    st.dataframe(tabela.style.format({"META R$":_m,"REALIZADO R$":_m,"PROJEÇÃO R$":_m,"GAP R$":_m,"% GAP":_p,"2025 R$":_m,"CRESC.":_p})
        .applymap(_cp,subset=["PRIORIDADE"]).applymap(_cn,subset=["GAP R$","CRESC."]),
        use_container_width=True,hide_index=True,height=480)
    st.divider()

    if st.button("🚀 Gerar roteiro do dia",key="r_rot"):
        rot=pd.concat([df[df["PRIORIDADE"]=="🔴 URGENTE"].sort_values(["CIDADE","GAP"]),
                       df[df["PRIORIDADE"]=="🟡 ATENÇÃO"].sort_values(["CIDADE","GAP"])]).head(10)
        if rot.empty: st.warning("Nenhum urgente ou em atenção.")
        else:
            st.success("Roteiro gerado!")
            for cid in rot["CIDADE"].dropna().unique():
                b=rot[rot["CIDADE"]==cid]; st.markdown(f"#### 📍 {cid}")
                for i,row in enumerate(b.itertuples(),1):
                    st.markdown(f"**{i}. {row.CLIENTE_FINAL}** · {row.PRIORIDADE} · GAP: {_m(row.GAP)} · {row.ACAO}")
    st.divider()

    res=df.groupby("CIDADE",dropna=False).agg(
        CLIENTES=("CLIENTE_FINAL","count"),URGENTES=("PRIORIDADE",lambda x:(x=="🔴 URGENTE").sum()),
        ATENCAO=("PRIORIDADE",lambda x:(x=="🟡 ATENÇÃO").sum()),
        META=("META","sum"),REALIZADO=("REALIZADO","sum"),PROJECAO=("PROJECAO","sum"),GAP=("GAP","sum"),
    ).reset_index().sort_values(["URGENTES","GAP"],ascending=[False,True])
    st.dataframe(res.style.format({"META":_m,"REALIZADO":_m,"PROJECAO":_m,"GAP":_m}).applymap(_cn,subset=["GAP"]),
        use_container_width=True,hide_index=True)
    st.divider()

    cli=st.selectbox("🎯 Análise individual",df["CLIENTE_FINAL"].astype(str).unique(),key="r_cli")
    d=df[df["CLIENTE_FINAL"].astype(str)==cli].iloc[0]
    r1,r2,r3,r4=st.columns(4)
    r1.metric("Meta",_m(d["META"])); r2.metric("Realizado",_m(d["REALIZADO"]))
    r3.metric("Projeção",_m(d["PROJECAO"])); r4.metric("GAP",_m(d["GAP"]))
    st.markdown(f"""<div class="section-card">
**{d["CLIENTE_FINAL"]}** · {d["BANDEIRA_FINAL"]} · {d.get("CIDADE","")}
**Prioridade:** {d["PRIORIDADE"]} · **Ação:** {d["ACAO"]}
---
💬 *"Meta {_m(d["META"])}, realizou {_m(d["REALIZADO"])}, projetando {_m(d["PROJECAO"])}. GAP de {_m(d["GAP"])} ({_p(d["PERC_GAP"])})."*
</div>""",unsafe_allow_html=True)
    primeiro=df.iloc[0]
    st.error(f"👉 Comece por: {primeiro['CLIENTE_FINAL']} · {primeiro['PRIORIDADE']} · {primeiro['ACAO']}")

# ══════════════════════════════════════════════════════════════
# COTABOT
# ══════════════════════════════════════════════════════════════
elif pagina=="cotabot":
    def norm(t):
        t=str(t).strip().lower()
        for k,v in {"ç":"c","ã":"a","á":"a","à":"a","â":"a","é":"e","ê":"e","í":"i","ó":"o","ô":"o","õ":"o","ú":"u"}.items(): t=t.replace(k,v)
        return t
    def leia(s): return s.astype(str).str.strip().str.replace(".0","",regex=False).str.replace(" ","",regex=False).str.replace("-","",regex=False)
    def conv(s):
        if pd.api.types.is_numeric_dtype(s): return pd.to_numeric(s,errors="coerce")
        s2=s.astype(str).str.strip().str.replace("R$","",regex=False).str.replace(" ","",regex=False)
        tv=s2.str.contains(",",regex=False,na=False)
        s2.loc[tv]=s2.loc[tv].str.replace(".",",",regex=False).str.replace(",",".",regex=False)
        return pd.to_numeric(s2,errors="coerce")
    def fmt(v):
        if pd.isna(v) or v=="": return ""
        try: return f"{float(v):.2f}".replace(".",",")
        except: return ""
    def xl(f,s=0):
        f.seek(0); n=f.name.lower()
        if n.endswith(".xlsx"): return pd.read_excel(f,engine="openpyxl",sheet_name=s,header=None)
        if n.endswith(".xls"):  return pd.read_excel(f,engine="xlrd",sheet_name=s,header=None)
        return pd.read_excel(f,sheet_name=s,header=None)
    def det_h(db,words,lim=25):
        bl,bs=0,-1
        for i in range(min(len(db),lim)):
            row=[norm(x) for x in db.iloc[i].fillna("").astype(str).tolist()]
            sc=sum(1 for c in row for w in words if w in c)
            if sc>bs: bs=sc; bl=i
        return bl
    def mk(db,h):
        cab=db.iloc[h].fillna("").astype(str).tolist()
        d=db.iloc[h+1:].copy().reset_index(drop=True); d.columns=cab; return d
    def fc(cols,names):
        mp={norm(c):c for c in cols}
        for n in names:
            nn=norm(n)
            for cn,co in mp.items():
                if nn==cn: return co
            for cn,co in mp.items():
                if nn in cn: return co
        return None
    def escrever_xlsx(f,aba,hrow,pcol,precos):
        f.seek(0); wb=load_workbook(f); ws=wb[aba]
        for i,p in enumerate(precos):
            ws.cell(row=hrow+2+i,column=pcol+1,value=float(p) if pd.notna(p) else None)
        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

    st.markdown('<div class="app-header"><h1>💰 CotaBot</h1><p>Cruzamento por EAN · Preenchimento automático · Download pronto</p></div>',unsafe_allow_html=True)
    st.markdown("#### 1. Base da empresa")
    base_f=st.file_uploader("Base de produtos (EAN + preço + estoque)",type=["xlsx","xls"],key="c_base")
    base_df=None
    if base_f:
        braw=xl(base_f); bh=det_h(braw,["codigo ean","código ean","descricao","laboratorio","st","preco nf","estoque"])
        base_df=mk(braw,bh); st.dataframe(base_df.head(8),use_container_width=True)

    st.markdown("---"); st.markdown("#### 2. Cotação do cliente")
    cot_f=st.file_uploader("Planilha de cotação",type=["xlsx","xls"],key="c_cot")
    cot_df=None; craw=None; aba=None; hrow=None
    if cot_f:
        if cot_f.name.lower().endswith(".xlsx"):
            cot_f.seek(0); wb2=load_workbook(cot_f,read_only=True); abas=wb2.sheetnames; wb2.close()
        else: abas=[0]
        aba=st.selectbox("Aba",abas,key="c_aba") if len(abas)>1 else abas[0]
        craw=xl(cot_f,sheet_name=aba)
        ch=det_h(craw,["ean","codigo ean","produto","descricao","qtd","preco","fabricante"])
        hrow=st.number_input("Linha do cabeçalho",min_value=1,max_value=max(1,len(craw)),value=int(ch+1),step=1,key="c_h")-1
        cot_df=mk(craw,hrow); st.dataframe(cot_df.head(8),use_container_width=True)

    if base_df is not None and cot_df is not None:
        st.markdown("---"); st.subheader("3. Associar colunas")
        s1,s2=st.columns(2)
        ob=["-- Selecionar --"]+list(base_df.columns); oc=["-- Selecionar --"]+list(cot_df.columns)
        be=fc(base_df.columns,["ean","codigo ean","código ean","gtin"])
        bp=fc(base_df.columns,["preço real","preco real","preço final","valor final"])
        bs=fc(base_df.columns,["st","valor st"]); bn=fc(base_df.columns,["preço nf","preco nf","valor nf"])
        bst=fc(base_df.columns,["estoque","saldo","disponivel"])
        ce=fc(cot_df.columns,["ean","codigo ean","gtin"]); cp2=fc(cot_df.columns,["preço","preco","valor","price"])
        with s1:
            st.markdown("**Base**")
            col_be=st.selectbox("EAN base",ob,index=ob.index(be) if be in ob else 0,key="c_be")
            modo=st.radio("Preço",["Usar PREÇO REAL","ST + PREÇO NF"],index=0 if bp else 1,key="c_modo")
            if modo=="Usar PREÇO REAL":
                col_bp=st.selectbox("PREÇO REAL",ob,index=ob.index(bp) if bp in ob else 0,key="c_bp"); col_bs=col_bn=None
            else:
                col_bs=st.selectbox("ST",ob,index=ob.index(bs) if bs in ob else 0,key="c_bs")
                col_bn=st.selectbox("PREÇO NF",ob,index=ob.index(bn) if bn in ob else 0,key="c_bn"); col_bp=None
            est_min=st.number_input("Estoque mínimo",min_value=0,value=1,step=1,key="c_est"); col_bst=bst
        with s2:
            st.markdown("**Cotação**")
            col_ce=st.selectbox("EAN cotação",oc,index=oc.index(ce) if ce in oc else 0,key="c_ce")
            col_cp=st.selectbox("PREÇO cotação",oc,index=oc.index(cp2) if cp2 in oc else 0,key="c_cp")
        st.markdown("---")
        if st.button("⚡ Processar cotação",use_container_width=True,key="c_proc"):
            try:
                bp3=base_df.copy(); cp3=cot_df.copy()
                bp3[col_be]=leia(bp3[col_be]); cp3[col_ce]=leia(cp3[col_ce])
                if col_bst: bp3[col_bst]=pd.to_numeric(bp3[col_bst],errors="coerce").fillna(0)
                if modo=="Usar PREÇO REAL": bp3["_P_"]=conv(bp3[col_bp])
                else:
                    bp3["_S_"]=conv(bp3[col_bs]); bp3["_N_"]=conv(bp3[col_bn])
                    bp3["_P_"]=bp3["_S_"].fillna(0)+bp3["_N_"].fillna(0)
                bf=bp3[bp3[col_bst]>=est_min].copy() if col_bst else bp3.copy()
                bm=bf[[col_be,"_P_"]].drop_duplicates(subset=[col_be])
                res=cp3.merge(bm,left_on=col_ce,right_on=col_be,how="left")
                nums=res["_P_"].tolist(); prev=[fmt(x) for x in nums]
                tot=len(cp3); enc=int(pd.notna(res["_P_"]).sum())
                a,b,c=st.columns(3); a.metric("Total",tot); b.metric("✅ Encontrados",enc); c.metric("❌ Não encontrados",tot-enc)
                pv=cp3.copy(); pv[col_cp]=pv[col_cp].astype("object"); pv[col_cp]=prev
                st.dataframe(pv.head(20),use_container_width=True)
                cab_row=craw.iloc[hrow].fillna("").astype(str).tolist()
                pidx=next((i for i,v in enumerate(cab_row) if str(v).strip()==str(col_cp).strip()),None)
                if pidx is None: st.error("Coluna de preço não localizada."); st.stop()
                if cot_f.name.lower().endswith(".xlsx"):
                    out=escrever_xlsx(cot_f,aba,hrow,pidx,nums)
                    st.download_button("⬇️ Baixar cotação preenchida",out,"cotacao_preenchida.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,key="c_dl")
                else:
                    rb=craw.copy(); rb[pidx]=rb[pidx].astype("object")
                    for i,p in enumerate(nums): rb.iat[hrow+1+i,pidx]=float(p) if pd.notna(p) else None
                    buf=io.BytesIO()
                    with pd.ExcelWriter(buf,engine="openpyxl") as w: rb.to_excel(w,index=False,header=False)
                    buf.seek(0); st.download_button("⬇️ Baixar",buf.getvalue(),"cotacao_preenchida.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,key="c_dl2")
            except Exception as e: st.error(f"Erro: {e}")
    else: st.info("Suba a base e a cotação para configurar.")

# ══════════════════════════════════════════════════════════════
# COBRANÇA / VISITAS (slots)
# ══════════════════════════════════════════════════════════════
elif pagina=="cobranca":
    st.markdown('<div class="app-header"><h1>🧾 Cobrança Inteligente</h1><p>Clientes vencidos e prioridade de cobrança.</p></div>',unsafe_allow_html=True)
    st.info("📨 Envie o código do app de Cobrança para ativar este módulo.")
    st.markdown("**Este módulo vai mostrar:**\n- ✅ Clientes com títulos vencidos\n- ✅ Dias em atraso e valor total\n- ✅ Prioridade automática de cobrança\n- ✅ Histórico e sugestão de abordagem")

elif pagina=="visitas":
    st.markdown('<div class="app-header"><h1>📍 Bússola de Visitas</h1><p>Rota otimizada e plano de ataque do dia.</p></div>',unsafe_allow_html=True)
    st.info("📨 Envie o código do app de Visitas para ativar este módulo.")
    st.markdown("**Este módulo vai mostrar:**\n- ✅ Rota otimizada por cidade\n- ✅ Clientes prioritários do dia\n- ✅ Mix faltante por cliente\n- ✅ Plano de ataque automático")